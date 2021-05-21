import openpyxl

from TestData.excelDemo import sheet


class HomePageData:
    test_HomePage_data = [{"fname": "Dinesh", "email": "din@abc.com", "gender": "Male"},
                          {"fname": "Kannan", "email": "kannan@abc.com", "gender": "Female"}]

    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\ptw\\Documents\\Test.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return[Dict]